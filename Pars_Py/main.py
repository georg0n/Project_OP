import openpyxl as op

filename = 'IIT_1-kurs_22_23_vesna_TANDEM_29.03.2023.xlsx'

wb = op.load_workbook(filename, data_only=True)

sheet = wb.active
max_column = sheet.max_column


group = input('Введите название группы: ')

print(group)

print('Понедельник \n---------------------------')
for i in range(2, max_column+1):
    search_group = sheet.cell(row=2, column = i).value
    numberOfColumn = i

    if not search_group:
        continue

    if (search_group == group):
        for j in range(4, 88):
            if j==17:
                print('-------------------- \n Вторник \n--------------------\n')
            elif j==31:
                print('--------------------\n Среда \n--------------------\n')
            elif j==45:
                print('-------------------\n Четверг \n--------------------\n')
            elif j==59:
                print('-------------------\n Пятница \n--------------------\n')
            elif j==73:
                print('-------------------\n Суббота \n--------------------\n')

            lesson = sheet.cell(row=j, column=i).value
            prepod = sheet.cell(row=j, column=i+2).value
            audit = sheet.cell(row=j, column=i+3).value
            time_start = sheet.cell(row=j, column=i-3).value
            time_end = sheet.cell(row=j, column=i-2).value

            if lesson:
                print(lesson, "|",  prepod, "|", audit)
        #print(lesson)
        #print(' ')

#17 31 45 59 73 87

import json
import sys

sys.stdout = ('declare.js')
