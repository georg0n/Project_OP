from flask import Flask, render_template, request, redirect, url_for, session, jsonify
import openpyxl as op
import os
import urllib.request


app = Flask(__name__) 


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        title = request.form.get('title')
        return redirect(url_for('start', title=title))
    return render_template('index.html')


@app.route('/start')
def start():
    title = request.args.get('title', None)

    if title[0] == 'И':
        if title[-1] == '2':
            filename = 'IIT_1-kurs_22_23_vesna_TANDEM_29.03.2023.xlsx'
            fl_download = 'https://webservices.mirea.ru/upload/iblock/145/nm0oup23k1406umgjq00mdewalurd4f6/IIT_1-kurs_22_23_vesna_27.04.2023.xlsx'
        if title[-1] == '1':
            filename = 'IIT_2-kurs_22_23_vesna_27.04.2023.xlsx'
            fl_download = 'https://webservices.mirea.ru/upload/iblock/a9a/vd0ew6sbdfgjp79zz3rtetmo84wy9qxd/IIT_2-kurs_22_23_vesna_15.05.2023.xlsx'
        if title[-1] == '0':
            filename = 'IIT_3-kurs_22_23_vesna_02.05.2023.xlsx'
            fl_download = 'https://webservices.mirea.ru/upload/iblock/b13/0005fcyi0poapgz9854iqrjuf7e1ywyf/IIT_3-kurs_22_23_vesna_02.05.2023.xlsx'
    elif title[0] == 'К':
        if title[-1] == '2':
            filename = 'III_1-kurs_22_23_vesna_10.05.2023.xlsx'
            fl_download = 'https://webservices.mirea.ru/upload/iblock/172/maekldms2d39o5ixvfljjf5bmertdxmy/III_1-kurs_22_23_vesna_11.05.2023.xlsx'
        if title[-1] == '1':
            filename = 'III_2-kurs_22_23_vesna_10.05.2023.xlsx'
            fl_download = 'https://webservices.mirea.ru/upload/iblock/1e2/lxmjye6l9kxtaztn9wn9wxp5rgec49ls/III_2-kurs_22_23_vesna_10.05.2023.xlsx'
        if title[-1] == '0':
            filename = 'III_3-kurs_22_23_vesna_28.04.2023.xlsx'
            fl_download = 'https://webservices.mirea.ru/upload/iblock/ba6/9ievf9qfdetjpex1dt76ey8yxcdu1gj7/III_3-kurs_22_23_vesna_28.04.2023.xlsx'
        if title[-1] == '9':
            filename = 'III_4-kurs_22_23_vesna_27.04.2023.xlsx'
            fl_download = 'https://webservices.mirea.ru/upload/iblock/e9d/mn77pkgk0pe2ti05hpe1lor1umm02qxs/III_4-kurs_22_23_vesna_27.04.2023.xlsx'
        if title[-1] == '8':
            filename = 'III_5-kurs_22_23_vesna_03.05.2023.xlsx'
            fl_download = 'https://webservices.mirea.ru/upload/iblock/837/f9gzrcggm5vd5y231ltsitrcs0r27ri8/III_5-kurs_22_23_vesna_03.05.2023.xlsx'
    elif title[0] == 'Б':
        if title[-1] == '2':
            filename = 'IKTST_1_k_vesna_22_23.xlsx'
            fl_download = 'https://webservices.mirea.ru/upload/iblock/e03/ohxuwbqgor2r6156fhra6zgpdalya35x/IKTST_1_k_vesna_22_23.xlsx'
        if title[-1] == '1':
            filename = 'IKTST_2_k_vesna_22_23.xlsx'
            fl_download = 'https://webservices.mirea.ru/upload/iblock/681/mdidmi1z0921ur8mon5u69ayfkcn3eej/IKTST_2_k_vesna_22_23.xlsx'
        if title[-1] == '0':
            filename = 'IKTST_3_k_vesna_22_23.xlsx'
            fl_download = 'https://webservices.mirea.ru/upload/iblock/b2b/d4atayz481r9xpknzn5iyays7rwf1mxp/IKTST_3_k_vesna_22_23.xlsx'
        if title[-1] == '9':
            filename = 'IKTST_4_k_vesna_22_23.xlsx'
            fl_download = 'https://webservices.mirea.ru/upload/iblock/dda/r2aanus36mbcbjm6j7ha5jxn2note5cr/IKTST_4_k_vesna_22_23.xlsx'
        if title[-1] == '8':
            filename = 'IKTST_5_k_vesna_22_23.xlsx'
            fl_download = 'https://webservices.mirea.ru/upload/iblock/b11/7txt2xnbtqxqavyphvqi98i8iuq1t0hr/IKTST_5_k_vesna_22_23.xlsx'
    elif title[0] == 'Т':
        if title[-1] == '2':
            filename = 'IPTIP_1-kurs_22_23_vesna_10.04.2023.xlsx'
            fl_download = 'https://webservices.mirea.ru/upload/iblock/204/l9p3e1uagan06chz6jysoboh3jge0ouj/IPTIP_1-kurs_22_23_vesna_10.04.2023.xlsx'
        if title[-1] == '1':
            filename = 'IPTIP_2-kurs_22_23_vesna_10.04.2023.xlsx'
            fl_download = 'https://webservices.mirea.ru/upload/iblock/9bf/2j5l9qh8ews8dzsc13cs6gyawbu86eut/IPTIP_2-kurs_22_23_vesna_15.05.2023.xlsx'
        if title[-1] == '0':
            filename = 'IPTIP_3-kurs_22_23_vesna_19.04.2023.xlsx'
            fl_download = 'https://webservices.mirea.ru/upload/iblock/83c/1du8odfzfxcivoof1hg3tm4mgopwm21j/IPTIP_3-kurs_22_23_vesna_19.04.2023.xlsx'
        if title[-1] == '9':
            filename = 'IPTIP_4-kurs_22_23_vesna_10.04.2023.xlsx'
            fl_download = 'https://webservices.mirea.ru/upload/iblock/f76/d0scyvzhr7fjeuwyatdxz8inqbq6yckc/IPTIP_4-kurs_22_23_vesna_10.04.2023.xlsx'
    elif title[0] == 'Р':
        if title[-1] == '2':
            filename = 'IRI_1-kurs_22_23_vesna_TANDEM_23.03.2023.xlsx'
            fl_download = 'https://webservices.mirea.ru/upload/iblock/8b7/ez2qa1lk4v8fcw4fdmr1o5rdmayrh8if/IRI_1-kurs_22_23_vesna_TANDEM_23.03.2023.xlsx'
        if title[-1] == '1':
            filename = 'IRI_2-kurs_22_23_vesna_TANDEM_22.03.2023.xlsx'
            fl_download = 'https://webservices.mirea.ru/upload/iblock/b94/ftsyi6m8r3i6mgpp3c7ndyegdca5vokl/IRI_2-kurs_22_23_vesna_TANDEM_22.03.2023.xlsx'
        if title[-1] == '0':
            filename = 'IRI_3-kurs_22_23_vesna_10.04.2023.xlsx'
            fl_download = 'https://webservices.mirea.ru/upload/iblock/62f/egi8fq6rr5le6zdjsgo3mk95pnba3rcr/IRI_3-kurs_22_23_vesna_10.04.2023.xlsx'
        if title[-1] == '9':
            filename = 'IRI_4-kurs_22_23_vesna_TANDEM_22.03.2023.xlsx'
            fl_download = 'https://webservices.mirea.ru/upload/iblock/676/pgag5yfxq6ok0lcffatx63y3nurjyngs/IRI_4-kurs_22_23_vesna_TANDEM_22.03.2023.xlsx'
        if title[-1] == '8':
            filename = 'IRI_5-kurs_22_23_vesna_TANDEM_22.03.2023.xlsx'
            fl_download = 'https://webservices.mirea.ru/upload/iblock/eb9/qmdqyco53me19uy33phniunngjps31ww/IRI_5-kurs_22_23_vesna_TANDEM_22.03.2023.xlsx'
    elif title[0] == 'У':
        if title[-1] == '2':
            filename = 'ITU_1-kurs_22_23_vesna_03.05.2023.xlsx'
            fl_download = 'https://webservices.mirea.ru/upload/iblock/a3c/kot4w9jwg732lhse6jfcinj08vf9ei77/ITU_1-kurs_22_23_vesna_03.05.2023.xlsx'
        if title[-1] == '1':
            filename = 'ITU_2-kurs_22_23_vesna_05.05.2023.xlsx'
            fl_download = 'https://webservices.mirea.ru/upload/iblock/b70/xsym3ytahczem8zdcmmxc87gxbrsjkl8/ITU_2-kurs_22_23_vesna_11.05.2023.xlsx'
        if title[-1] == '0':
            filename = 'ITU_3-kurs_22_23_vesna_02.05.2023.xlsx'
            fl_download = 'https://webservices.mirea.ru/upload/iblock/b52/nruiz2x0qr7ozs0euld1eat6bj0ur6xo/ITU_3-kurs_22_23_vesna_11.05.2023.xlsx'


    url = fl_download
    urllib.request.urlretrieve(url, filename)


    wb = op.load_workbook(filename, data_only=True)

    sheet = wb.active
    max_column = sheet.max_column
    arr = []
    arr_prep = []
    arr_surname = []
    arr_audit = []
    group = title

    for i in range(2, max_column+1):
        search_group = sheet.cell(row=2, column=i).value
        numberOfColumn = i
        if not search_group:
            continue

        if (search_group == group):
            for j in range(4, 88):
                prep = sheet.cell(row=j, column=i+1).value
                arr.append(sheet.cell(row=j, column=i).value)
                surname = sheet.cell(row=j, column=i+2).value
                audit = sheet.cell(row=j, column=i+3).value

                if prep != "":
                    arr_prep.append(" || " + " (" + (prep) + ") " + " || ")
                else:
                    arr_prep.append("")

                if surname != "":
                    arr_surname.append(" (" + (surname) + ") " + " || ")
                else:
                    arr_surname.append("")

                if audit != "":
                    arr_audit.append("(" + (audit) + ")")
                else:
                    arr_audit.append("")
    os.remove(filename)

    return render_template('start.html', title=title, arr=arr, arr_prep=arr_prep, arr_surname=arr_surname, arr_audit=arr_audit)


if __name__ == '__main__':
    app.run(debug=True)












